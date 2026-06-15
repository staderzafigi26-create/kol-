#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime, time, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
REPORT_DIR = ROOT / "data" / "reports"
MARKETING_FILE = Path("/Users/ryan/Downloads/Yozma-红人营销总表 (5).xlsx")
CONTRACT_FILE = Path("/Users/ryan/Downloads/Yozma-红人合同签署表 (2).xlsx")
ORDER_FILES = [
    ("UK", Path("/Users/ryan/Downloads/orders-08-Jun-2026-14-Jun-2026.csv")),
    ("US", Path("/Users/ryan/Downloads/orders-08-Jun-2026-14-Jun-2026 (1).csv")),
    ("EU", Path("/Users/ryan/Downloads/orders-08-Jun-2026-14-Jun-2026 (2).csv")),
    ("CA", Path("/Users/ryan/Downloads/orders-08-Jun-2026-14-Jun-2026 (3).csv")),
]
REGION_ORDER = ["UK", "US", "EU", "CA"]
START = date(2026, 6, 8)
END = date(2026, 6, 14)
TZ = timezone(timedelta(hours=8))


LIVE_SALES_WORDS = ("直播", "带货", "tiktok shop", "tk shop", "shopify collabs")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def norm_key(value: Any) -> str:
    text = clean(value).lower()
    text = re.sub(r"https?://", "", text)
    text = re.sub(r"^www\.", "", text)
    text = re.sub(r"[@\s/_\\.-]+", "", text)
    return text


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            pass
    match = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    match = re.search(r"(\d{1,2})/(\d{1,2})", text)
    if match:
        return date(2026, int(match.group(1)), int(match.group(2)))
    return None


def parse_datetime_local(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        text = clean(value)
        if not text:
            return None
        try:
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
                try:
                    dt = datetime.strptime(text[:19], fmt)
                    break
                except ValueError:
                    dt = None
            if dt is None:
                return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TZ)
    return dt.astimezone(TZ)


def in_period(value: Any) -> bool:
    d = parse_date(value)
    return d is not None and START <= d <= END


def parse_num(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = clean(value).replace(",", "")
    if not text:
        return 0.0
    multiplier = 1.0
    lower = text.lower()
    if "万" in lower:
        multiplier = 10000.0
    elif "k" in lower:
        multiplier = 1000.0
    elif "m" in lower:
        multiplier = 1000000.0
    match = re.search(r"-?\d+(?:\.\d+)?", lower)
    return float(match.group(0)) * multiplier if match else 0.0


def format_int(value: float) -> str:
    return str(int(round(value)))


def format_money_by_currency(counter: dict[str, float]) -> str:
    parts = []
    for currency, amount in sorted(counter.items()):
        if abs(amount) >= 1000:
            parts.append(f"{currency} {amount:,.0f}")
        else:
            parts.append(f"{currency} {amount:,.2f}")
    return "；".join(parts) if parts else "待补充"


def is_live_or_sales(row: dict[str, Any]) -> bool:
    text = " ".join(clean(v).lower() for v in row.values())
    return any(word in text for word in LIVE_SALES_WORDS)


def read_xlsx_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [clean(v) for v in values[0]]
    rows = []
    for raw in values[1:]:
        if not any(clean(v) for v in raw):
            continue
        row = {headers[i]: raw[i] if i < len(raw) else "" for i in range(len(headers)) if headers[i]}
        rows.append(row)
    return rows


def read_json_array(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8") as fh:
        data = json.load(fh)
    return data if isinstance(data, list) else []


def field(record: dict[str, Any], name: str, default: Any = "") -> Any:
    return record.get("fields", {}).get(name, default)


def platform_label(value: Any) -> str:
    p = clean(value).lower().replace(" ", "")
    if "youtube" in p and "short" in p:
        return "YouTube Shorts"
    if "youtube" in p:
        return "YouTube Video"
    if "instagram" in p or "reels" in p:
        return "Instagram Reels"
    if "tiktok" in p or p == "tk":
        return "TikTok"
    return clean(value) or "未标注"


def extract_url(value: Any) -> str:
    if isinstance(value, dict):
        return clean(value.get("link") or value.get("text"))
    return clean(value)


def build_influencer_index(influencers: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    index = {}
    for item in influencers:
        f = item.get("fields", {})
        keys = [
            f.get("红人名称"),
            f.get("名字"),
            f.get("红人编码"),
            extract_url(f.get("红人链接")),
            extract_url(f.get("主平台链接（Instagram）")),
        ]
        for key in keys:
            nk = norm_key(key)
            if nk and nk not in index:
                index[nk] = f
    return index


def lookup_creator(index: dict[str, dict[str, Any]], creator: Any) -> dict[str, Any]:
    return index.get(norm_key(creator), {})


def summarize_counter(rows: list[dict[str, Any]], key: str) -> Counter:
    c = Counter()
    for row in rows:
        value = clean(row.get(key)) or "未标注"
        c[value] += 1
    return c


def make_marketing_rows() -> list[dict[str, Any]]:
    rows = read_xlsx_rows(MARKETING_FILE, "合作名单")
    output = []
    for row in rows:
        if not in_period(row.get("合作日期")):
            continue
        if is_live_or_sales(row):
            continue
        region = clean(row.get("地区")) or clean(row.get("红人编码")).split("-")[0]
        if region not in REGION_ORDER:
            region = region or "未标注"
        progress = clean(row.get("合作进度"))
        if progress in REGION_ORDER:
            progress = clean(row.get("合作模式")) or "已签合同"
        output.append({
            "合作日期": parse_date(row.get("合作日期")).isoformat(),
            "负责人": clean(row.get("负责人")) or "未标注",
            "地区": region,
            "红人编码": clean(row.get("红人编码")),
            "红人名称": clean(row.get("名字")),
            "类型": clean(row.get("类型")),
            "粉丝": clean(row.get("粉丝")),
            "量级": clean(row.get("量级")),
            "合作进度": progress,
            "合作模式": clean(row.get("合作模式")),
            "合作车型": clean(row.get("合作车型")),
            "合作周期": clean(row.get("合作周期")),
            "联盟Code": clean(row.get("联盟Code")),
            "Instagram": clean(row.get("主平台链接（Instagram）")),
            "TikTok": clean(row.get("主平台链接（TikTok）")),
            "YouTube": clean(row.get("主平台链接（Youtube）") or row.get("主平台链接（YouTube）")),
            "预计交付": clean(row.get("合作数量") or row.get("其他交付物") or row.get("合作周期")),
        })
    return output


def make_contract_rows() -> list[dict[str, Any]]:
    output = []
    for sheet in ("红人合作合同信息", "带货红人合同信息"):
        for row in read_xlsx_rows(CONTRACT_FILE, sheet):
            if not in_period(row.get("盖章日期")):
                continue
            if sheet == "带货红人合同信息" or is_live_or_sales(row):
                continue
            reason = clean(row.get("签署事由"))
            output.append({
                "盖章日期": parse_date(row.get("盖章日期")).isoformat(),
                "红人名称": clean(row.get("红人名字")),
                "涉及区域": clean(row.get("涉及区域")),
                "红人类型": clean(row.get("红人类型")),
                "合同类型": clean(row.get("合同类型")),
                "合同金额": clean(row.get("合同涉及总额")),
                "币种": clean(row.get("币种")),
                "交付摘要": reason[:260],
                "原始签署事由": reason,
            })
    return output


def make_video_rows(influencer_index: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = read_json_array(ROOT / "data" / "local" / "videos.json")
    output = []
    start_dt = datetime.combine(START, time.min, tzinfo=TZ)
    end_dt = datetime.combine(END + timedelta(days=1), time.min, tzinfo=TZ)
    for item in rows:
        f = item.get("fields", {})
        dt = parse_datetime_local(f.get("timestamp") or f.get("上线时间"))
        if not dt or not (start_dt <= dt < end_dt):
            continue
        creator = clean(f.get("红人名称") or f.get("creator"))
        matched = lookup_creator(influencer_index, creator)
        views = parse_num(f.get("7日成熟声量") or f.get("videoPlayCount") or f.get("videoViewCount"))
        mature30 = parse_num(f.get("30日成熟声量"))
        output.append({
            "上线日期": dt.strftime("%Y-%m-%d %H:%M"),
            "红人名称": creator,
            "负责人": clean(f.get("负责人") or matched.get("负责人")) or "未标注",
            "地区": clean(f.get("地区") or matched.get("地区") or clean(matched.get("红人编码")).split("-")[0]) or "未标注",
            "平台": platform_label(f.get("平台")),
            "7日成熟声量/当前播放": int(round(views)),
            "30日成熟声量": int(round(mature30)) if mature30 else "",
            "互动率": clean(f.get("互动率") or f.get("engagementRate")),
            "链接": extract_url(f.get("url") or f.get("videoUrl")),
            "是否计入合同交付": clean(f.get("是否计入合同交付")),
        })
    output.sort(key=lambda r: r["7日成熟声量/当前播放"], reverse=True)
    return output


def read_order_file(region: str, path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        output = []
        for row in reader:
            dt = parse_datetime_local(row.get("Order Date"))
            if not dt or not (START <= dt.date() <= END):
                continue
            output.append({
                "市场": region,
                "订单时间": dt.strftime("%Y-%m-%d %H:%M"),
                "订单号": clean(row.get("Order Number")),
                "Order ID": clean(row.get("Order ID")),
                "达人": clean(row.get("Affiliate Name")),
                "Affiliate Email": clean(row.get("Affiliate Email")),
                "Discount codes": clean(row.get("Discount codes")),
                "订单金额": parse_num(row.get("Order Total")),
                "订单小计": parse_num(row.get("Order Subtotal")),
                "佣金": parse_num(row.get("Affiliate Commission")),
                "币种": clean(row.get("Currency")),
                "订单状态": clean(row.get("Order Status")),
                "Conversion Source": clean(row.get("Conversion Source")),
            })
    return output


def make_order_rows() -> tuple[list[dict[str, Any]], dict[str, int]]:
    seen = {}
    duplicates = Counter()
    for region, path in ORDER_FILES:
        for row in read_order_file(region, path):
            key = row.get("Order ID") or row.get("订单号") or "|".join(
                [row.get("订单时间", ""), row.get("Affiliate Email", ""), str(row.get("订单金额", ""))]
            )
            if key in seen:
                duplicates[region] += 1
                continue
            seen[key] = row
    rows = list(seen.values())
    rows.sort(key=lambda r: (REGION_ORDER.index(r["市场"]) if r["市场"] in REGION_ORDER else 99, r["订单时间"]))
    return rows, dict(duplicates)


def summarize_orders(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary = []
    for region in REGION_ORDER:
        subset = [r for r in rows if r["市场"] == region]
        revenue = defaultdict(float)
        subtotal = defaultdict(float)
        commission = defaultdict(float)
        for row in subset:
            currency = row["币种"] or region
            revenue[currency] += row["订单金额"]
            subtotal[currency] += row["订单小计"]
            commission[currency] += row["佣金"]
        summary.append({
            "市场": region,
            "订单数": len(subset),
            "Revenue/Order Total": format_money_by_currency(revenue),
            "Gross Sales/Subtotal": format_money_by_currency(subtotal),
            "Commission": format_money_by_currency(commission),
            "Top Affiliate": Counter(r["达人"] or "未标注" for r in subset).most_common(1)[0][0] if subset else "待补充",
        })
    return summary


def by_region(rows: list[dict[str, Any]], value_key: str | None = None) -> list[dict[str, Any]]:
    out = []
    for region in REGION_ORDER + ["未标注"]:
        subset = [r for r in rows if clean(r.get("地区") or r.get("涉及区域")) == region]
        if not subset:
            continue
        row = {"地区": region, "数量": len(subset)}
        if value_key:
            row[value_key] = sum(parse_num(r.get(value_key)) for r in subset)
        out.append(row)
    return out


def make_ryan_highlights(marketing: list[dict[str, Any]], contracts: list[dict[str, Any]], videos: list[dict[str, Any]]) -> list[dict[str, Any]]:
    video_by_creator = defaultdict(list)
    for video in videos:
        video_by_creator[norm_key(video["红人名称"])].append(video)
    contract_names = {norm_key(r["红人名称"]): r for r in contracts}
    output = []
    for row in marketing:
        if clean(row.get("负责人")).lower() != "ryan":
            continue
        key = norm_key(row.get("红人名称"))
        related = video_by_creator.get(key, [])
        best_views = max((parse_num(v.get("7日成熟声量/当前播放")) for v in related), default=0)
        contract = contract_names.get(key, {})
        reason = []
        if parse_num(row.get("粉丝")) >= 100000:
            reason.append("粉丝量级较高")
        if related:
            reason.append(f"本周已上线 {len(related)} 条")
        if best_views >= 50000:
            reason.append(f"最高声量 {format_int(best_views)}")
        if contract:
            reason.append("本周合同已盖章")
        output.append({
            "红人名称": row.get("红人名称"),
            "地区": row.get("地区"),
            "合作日期": row.get("合作日期"),
            "合作进度": row.get("合作进度"),
            "合作车型": row.get("合作车型"),
            "合作模式": row.get("合作模式"),
            "预计交付": row.get("预计交付") or contract.get("交付摘要", "")[:80],
            "本周上线条数": len(related),
            "最高7日声量/当前播放": int(best_views),
            "亮点判断": "；".join(reason) if reason else "新合作，建议继续跟进交付节点",
        })
    output.sort(key=lambda r: (r["最高7日声量/当前播放"], r["本周上线条数"]), reverse=True)
    return output


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["暂无数据"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    header_fill = PatternFill("solid", fgColor="172033")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        max_len = min(max(len(clean(header)), *(len(clean(r.get(header))) for r in rows[:200])) + 2, 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_markdown(summary: dict[str, Any], path: Path) -> None:
    lines = [
        "# Yozma KOL 周报数据包（2026-06-08 至 2026-06-14）",
        "",
        "## 核心结论",
        f"- 本周新增/更新合作红人：{summary['cooperation_count']} 位，Ryan 负责：{summary['ryan_count']} 位。",
        f"- 本周上线视频：{summary['video_count']} 条，参与达人：{summary['video_creator_count']} 位，7日成熟声量/当前播放合计：{summary['video_views']:,}。",
        f"- 本周订单去重后：{summary['order_count']} 单；重复订单已跳过：{summary['duplicate_count']} 单。",
        "",
        "## 合作数据",
        summary["cooperation_text"],
        "",
        "## 上线数据",
        summary["video_text"],
        "",
        "## 出单数据",
        summary["order_text"],
        "",
        "## 重点红人合作进度",
        summary["highlight_text"],
        "",
        "## 数据口径",
        "- 周期：2026-06-08 00:00 至 2026-06-14 23:59，按北京时间/本地时间口径。",
        "- 订单：四个订单 CSV 按 UK、US、EU、CA 顺序导入；优先使用 Order ID 去重，缺失时使用订单号兜底。",
        "- 直播/带货红人：已从本次合作与合同统计中剔除。",
        "- 多币种收入不强行折算美元，按原币种分别汇总。",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    influencers = read_json_array(ROOT / "data" / "local" / "influencers.json")
    influencer_index = build_influencer_index(influencers)
    marketing = make_marketing_rows()
    contracts = make_contract_rows()
    videos = make_video_rows(influencer_index)
    orders, duplicate_by_region = make_order_rows()
    ryan_highlights = make_ryan_highlights(marketing, contracts, videos)

    cooperation_by_region = []
    for region in REGION_ORDER:
        subset = [r for r in marketing if r["地区"] == region]
        cooperation_by_region.append({
            "地区": region,
            "合作红人数": len(subset),
            "Ryan负责": sum(1 for r in subset if r["负责人"].lower() == "ryan"),
            "主要进度": "；".join(f"{k} {v}" for k, v in Counter(r["合作进度"] or "未标注" for r in subset).most_common(4)) or "待补充",
            "主要负责人": "；".join(f"{k} {v}" for k, v in Counter(r["负责人"] for r in subset).most_common(4)) or "待补充",
        })

    video_summary_region = []
    for region in REGION_ORDER + ["未标注"]:
        subset = [v for v in videos if v["地区"] == region]
        if subset:
            video_summary_region.append({
                "地区": region,
                "上线视频": len(subset),
                "上线达人": len({norm_key(v["红人名称"]) for v in subset}),
                "7日成熟声量/当前播放": sum(parse_num(v["7日成熟声量/当前播放"]) for v in subset),
                "Top视频达人": subset[0]["红人名称"],
                "Top视频平台": subset[0]["平台"],
            })

    video_summary_platform = []
    for platform, count in Counter(v["平台"] for v in videos).most_common():
        subset = [v for v in videos if v["平台"] == platform]
        video_summary_platform.append({
            "平台": platform,
            "上线视频": count,
            "上线达人": len({norm_key(v["红人名称"]) for v in subset}),
            "7日成熟声量/当前播放": sum(parse_num(v["7日成熟声量/当前播放"]) for v in subset),
            "均播": int(sum(parse_num(v["7日成熟声量/当前播放"]) for v in subset) / count) if count else 0,
            "Top达人": subset[0]["红人名称"] if subset else "",
        })

    order_summary = summarize_orders(orders)
    owner_summary = []
    for owner, count in summarize_counter(marketing, "负责人").most_common():
        owner_summary.append({"负责人": owner, "合作红人数": count})

    overview = [
        {"指标": "周期", "数值": "2026-06-08 至 2026-06-14", "说明": "本次周报周期"},
        {"指标": "合作红人", "数值": len(marketing), "说明": "营销总表合作日期落在周期内，已剔除直播/带货"},
        {"指标": "合同记录", "数值": len(contracts), "说明": "合同盖章日期落在周期内，已剔除直播/带货"},
        {"指标": "Ryan 本周合作", "数值": sum(1 for r in marketing if r["负责人"].lower() == "ryan"), "说明": "用于重点红人合作进度"},
        {"指标": "上线视频", "数值": len(videos), "说明": "本地视频库 timestamp 落在周期内"},
        {"指标": "上线达人", "数值": len({norm_key(v["红人名称"]) for v in videos}), "说明": "按红人名称去重"},
        {"指标": "7日成熟声量/当前播放", "数值": sum(parse_num(v["7日成熟声量/当前播放"]) for v in videos), "说明": "有7日成熟声量优先使用，否则用当前播放"},
        {"指标": "订单数", "数值": len(orders), "说明": "四个订单表按 Order ID/订单号去重"},
        {"指标": "跳过重复订单", "数值": sum(duplicate_by_region.values()), "说明": str(duplicate_by_region) if duplicate_by_region else "无重复"},
    ]

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "核心总览", overview)
    add_sheet(wb, "合作-区域汇总", cooperation_by_region)
    add_sheet(wb, "合作-负责人汇总", owner_summary)
    add_sheet(wb, "合作-明细", marketing)
    add_sheet(wb, "合同-明细", contracts)
    add_sheet(wb, "上线-区域汇总", video_summary_region)
    add_sheet(wb, "上线-平台汇总", video_summary_platform)
    add_sheet(wb, "上线-视频明细", videos)
    add_sheet(wb, "出单-区域汇总", order_summary)
    add_sheet(wb, "出单-订单明细", orders)
    add_sheet(wb, "重点-Ryan合作进度", ryan_highlights)

    xlsx_path = REPORT_DIR / "Yozma-KOL周报-2026-06-08_2026-06-14.xlsx"
    md_path = REPORT_DIR / "Yozma-KOL周报-2026-06-08_2026-06-14.md"
    wb.save(xlsx_path)

    top_regions = "；".join(f"{r['地区']} {r['合作红人数']}位" for r in cooperation_by_region if r["合作红人数"])
    top_platforms = "；".join(f"{r['平台']} {r['上线视频']}条 / {int(r['7日成熟声量/当前播放']):,}播放" for r in video_summary_platform[:4])
    order_text = "；".join(f"{r['市场']} {r['订单数']}单，{r['Revenue/Order Total']}" for r in order_summary)
    highlight_text = "；".join(
        f"{r['红人名称']}（{r['地区']}，{r['合作进度']}，{r['亮点判断']}）" for r in ryan_highlights[:6]
    ) or "本周 Ryan 新合作中暂无可自动识别的高表现上线视频，建议继续跟进合同交付节点。"
    summary = {
        "cooperation_count": len(marketing),
        "ryan_count": sum(1 for r in marketing if r["负责人"].lower() == "ryan"),
        "video_count": len(videos),
        "video_creator_count": len({norm_key(v["红人名称"]) for v in videos}),
        "video_views": int(sum(parse_num(v["7日成熟声量/当前播放"]) for v in videos)),
        "order_count": len(orders),
        "duplicate_count": sum(duplicate_by_region.values()),
        "cooperation_text": f"本周合作主要分布：{top_regions or '待补充'}。负责人维度以 " + "、".join(f"{r['负责人']} {r['合作红人数']}位" for r in owner_summary[:5]) + " 为主。",
        "video_text": f"本周上线表现按平台看：{top_platforms or '待补充'}。播放口径优先取7日成熟声量，未成熟视频使用当前播放作临时观察。",
        "order_text": f"按 UK、US、EU、CA 顺序汇总：{order_text}。多币种未做汇率折算。",
        "highlight_text": highlight_text,
    }
    build_markdown(summary, md_path)

    manifest = {
        "xlsx": str(xlsx_path),
        "markdown": str(md_path),
        "download_xlsx": f"/api/local/reports/{xlsx_path.name}",
        "download_markdown": f"/api/local/reports/{md_path.name}",
        "summary": summary,
        "sheets": [ws.title for ws in wb.worksheets],
    }
    manifest_path = REPORT_DIR / "Yozma-KOL周报-2026-06-08_2026-06-14.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
